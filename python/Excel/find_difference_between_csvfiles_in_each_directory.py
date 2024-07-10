import openpyxl
import os
from openpyxl import load_workbook
import pandas as pd
import math



def get_gb_xlsx_list():
    os.chdir("excel file main directory")
    file_list = os.listdir()
    file_list = [file for file in file_list if file.endswith(".xlsx") if file.startswith("DB")]
    file_list.remove('DB_EVENT_LIST.xlsx') # 한일에 없는 파일
    return file_list


# DB_GB 에 있는 엑셀파일만 비교하면 됨 다른건 한일거랑 같으니까
# DB_GB 에 있는 엑셀파일에 있는 시트들 가져와서 그대로 csv 오픈하기
def get_sheet_list():

    sheets = []
    gbxlsx = get_gb_xlsx_list()

    for i in gbxlsx:
        print(i + "에 있는 시트목록 추출중......")
        wb = openpyxl.load_workbook(filename=gb_dir + '\\' + i)
        sheetlist = wb.sheetnames
        if sheetlist[0][0] == "#":  # #define 제거
            for i in sheetlist[1:]:
                sheets.append(i)
        else :
            for j in sheetlist:
                sheets.append(j)
    print("추출완료")
    return sheets


def get_csv_from_kr(csvname):
    os.chdir(kr_csv_dir)
    prc = pd.read_csv(kr_csv_dir+"\\"+csvname+".csv" , engine='python')
    prc['ver'] = 'kr'
    return(prc)


def get_csv_from_gb(csvname):
    os.chdir(gb_csv_dir)
    prc = pd.read_csv(gb_csv_dir+"\\"+csvname+".csv" , engine='python')
    prc['ver'] = 'gb'
    return(prc)


def get_first_column(csvname):
    prc = pd.read_csv(kr_csv_dir + "\\" + csvname + ".csv" , engine='python')
    return prc.columns[0]



def find_changes_btw_kr_gb(csv_list):
    diff_count = 1
    first_sheet = 0

    for k in csv_list:

        kr = get_csv_from_kr(k)
        gb = get_csv_from_gb(k)
        col = get_first_column(k)

       
    # 두 데이터 프레임 합침

        df_concat = pd.concat([kr, gb], ignore_index=True)

        # 모든 컬럼 중복되는 데이터 삭제
        changes = df_concat.drop_duplicates(df_concat.columns[:-1], keep= 'last')

        duplicate_list = changes[changes[col].duplicated()][col].to_list()
        df_changed = changes[changes[col].isin(duplicate_list)]

        df_changed_gb = df_changed[df_changed['ver']=='gb'].iloc[:,:-1]
        df_changed_gb.sort_values(by=col,inplace=True)

        df_changed_kr =df_changed[df_changed['ver']=='kr'].iloc[:,:-1]
        df_changed_kr.sort_values(by=col,inplace=True)

        df_info_changes=df_changed_kr.copy()


        print(k + '.csv diff 중.......' + (str(diff_count) + '/' + str(len(csv_list))))


        for i in range(len(df_changed_gb.index)):

            for j in range(len(df_changed_gb.columns)):

                if str(df_changed_gb.iloc[i,j]) == 'nan' :
                    pass

                elif(df_changed_gb.iloc[i,j] != df_changed_kr.iloc[i,j]):
#                        print(df_changed_gb.iloc[i,j])
#                        print(df_changed_kr.iloc[i,j])

                    df_info_changes.iloc[i,j] = 'kr : ' + str(df_changed_kr.iloc[i,j]) + "==>" + ' gb : ' + str(df_changed_gb.iloc[i,j])


        if len(df_info_changes) == 0:  # 바뀐게 없으면 시트 안만듬
            pass
        else:

            os.chdir('directory to save difference')
            if first_sheet == 0 :

                with pd.ExcelWriter('diff_result_kr_to_gb.xlsx', mode='w', engine='openpyxl' ) as writer:
                   df_info_changes.to_excel(writer, sheet_name = k, index =False)

                first_sheet += 1

            else :

                with pd.ExcelWriter('diff_result_kr_to_gb.xlsx', mode='a', engine='openpyxl') as writer:
                   df_info_changes.to_excel(writer, sheet_name = k, index =False)


        diff_count += 1


def check_csv_in_kr(sheet_list) :

    removesheet = []
    os.chdir(kr_csv_dir)
    list_in_krcsv = os.listdir()
    for i in sheet_list:

        if i +'.csv' in list_in_krcsv:
            removesheet.append(i)

    return(removesheet)




if __name__ == '__main__' :

    kr_dir = "excel file directory to compare with main"
    gb_dir = "excel file main directory"
    kr_csv_dir = "csv file(= sheets in excel file) directory to compare with main"
    gb_csv_dir = "csv file(= sheets in excel file) main directory "

    gsl = get_sheet_list() # gb_dir에 있는 엑셀파일에 있는 시트 목록 모두 가져오기
    remove_no_csv_in_kr = check_csv_in_kr(gsl) # gb_csv에만 있는 몇몇 시트가 있어서 제거

    find_changes_btw_kr_gb(remove_no_csv_in_kr)
