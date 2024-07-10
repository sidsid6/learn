import openpyxl
import os
from openpyxl import load_workbook

def find_sheet(csvname):
    if csvname == 'db_quest':
        print("모든 퀘스트 디비를 찾아보세요")
    else :

        for i in file_list:
            print(i + " 찾는중")
            checkbreak = True

            # 읽어오려는 엑셀파일명에 디렉토리 + 파일명 구조로 입력
            wb = openpyxl.load_workbook(filename=chdir + i)
            sheetnames = wb.sheetnames
            for j in sheetnames:

                # csvname을 소문자로 입력받아서 소문자로 변경
                if j.lower() == csvname:
                    print(chdir + i + " 에 있습니다.")
                    checkbreak = False
                    break
                else:
                    continue

            if checkbreak == False:
                break


if __name__ == '__main__':

    os.chdir('working directory') # 입력받으면 좋을텐데 내가 쓸거니까 걍 고정으로 사용
    chdir = "working directory\\"
    
    # 찾고자 하는 디렉토리에 있는 엑셀 파일리스트 추출
    file_list = os.listdir()

    # 엑셀파일만 가져오기
    file_list = [file for file in file_list if file.endswith(".xlsx")]

    want_sheet_name = input("찾으려는 csv 이름을 입력하세요 (.csv 제외): ")
    find_sheet(want_sheet_name.lower())
