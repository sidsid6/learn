# 1. pandas  (dataframe 형식으로 변환하여 활용 / 대용량 데이터 처리 한정으로 빠름)

import pandas as pd

# 엑셀 파일 읽기
file_name = "test.xlsx"
df = pd.read_excel(file_name)

# 시트 이름들 출력하기
pd.read_excel(file_name, sheet_name = None)

# 특정 시트 불러오기
sheet1 = pd.read_excel(file_name, sheet_name = "test1")



ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ


# 2. openpyxl (빠르다)

from openpyxl import load_workbook

# 엑셀 파일 읽기
file_name = "test.xlsx"
lwb = load_workbook(file_name)

# 시트 이름들 불러오기
sheetnames = lwb.sheetnames

# 특정 시트 불러오기
sheet1 = lwb.get_sheet_by_name("test1")

sheet2 = lwb["test1"]
